import os
import re
import sys
import time
import json  # Добавили для работы с настройками
import subprocess
import importlib
from datetime import date, datetime

# ==========================================
# КОНСТАНТЫ ОФОРМЛЕНИЯ И НАСТРОЕК
# ==========================================
BOLD = "\033[1m"
RESET = "\033[0m"

# Определяем путь к папке, где лежит сам скрипт
script_dir = os.path.dirname(os.path.abspath(__file__))
# Полный путь к конфигу (чтобы он создавался рядом со скриптом)
CONFIG_FILE = os.path.join(script_dir, "config.json")
SORTING_SHEET_FILE = os.path.join(script_dir, "Sorting sheet.xlsx")
REQUIREMENTS_FILE = os.path.join(script_dir, "requirements.txt")

# ==========================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ (Утилиты)
# ==========================================

def print_error(message):
    """Вывод ошибки с красным восклицательным знаком."""
    print(f"❗️ {message}")


def ensure_dependencies():
    """Проверяет зависимости и при необходимости ставит их из requirements.txt."""
    required_modules = ("PyPDF2", "openpyxl")
    missing_modules = []

    for module_name in required_modules:
        try:
            importlib.import_module(module_name)
        except ImportError:
            missing_modules.append(module_name)

    if not missing_modules:
        return True

    print(f"ℹ️  Не найдены модули: {', '.join(missing_modules)}")
    print("ℹ️  Пробую установить зависимости из requirements.txt ...")

    if not os.path.exists(REQUIREMENTS_FILE):
        print_error("Файл requirements.txt не найден рядом со скриптом.")
        return False

    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "-r", REQUIREMENTS_FILE])
    except Exception as e:
        print_error(f"Не удалось установить зависимости из requirements.txt: {e}")
        print("ℹ️  Пробую установить недостающие модули напрямую ...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", *missing_modules])
        except Exception as direct_e:
            print_error(f"Не удалось установить зависимости автоматически: {direct_e}")
            print_error("Установите вручную: python -m pip install -r requirements.txt")
            return False

    # Проверяем, что установка действительно прошла успешно.
    for module_name in required_modules:
        try:
            importlib.import_module(module_name)
        except ImportError:
            print_error(f"Модуль {module_name} не установлен после автоустановки.")
            return False

    print("✅ Зависимости успешно установлены.")
    return True


if not ensure_dependencies():
    print("\nНажмите Enter для выхода...")
    input()
    sys.exit(1)

from PyPDF2 import PdfMerger


def load_config():
    """Загружает настройки из JSON файла."""
    if not os.path.exists(CONFIG_FILE):
        return None
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print_error(f"Ошибка чтения конфига: {e}")
        return None


def save_config(source_path, save_path):
    """Сохраняет пути в JSON файл."""
    data = {
        "source_path": source_path,
        "save_path": save_path
    }
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        # print(f"{BOLD}✔ Настройки путей сохранены.{RESET}") # Можно раскомментировать для отладки
    except Exception as e:
        print_error(f"Не удалось сохранить настройки: {e}")


def get_clean_path(prompt_text, allow_menu_codes=False):
    """
    Запрашивает путь.
    Текст запроса (prompt_text) делается жирным.
    """
    full_prompt = f"{BOLD}{prompt_text}:{RESET} "
    path = input(full_prompt).strip()

    if allow_menu_codes and path in ['0', '1', '9']:  # Добавили 9 для сброса
        return path

    if (path.startswith('"') and path.endswith('"')) or (path.startswith("'") and path.endswith("'")):
        path = path[1:-1]
    return path


def parse_folder_range(range_str):
    """Парсит строку диапазона (например, '3550-3553,3560')."""
    ranges = range_str.split(',')
    folder_numbers = set()
    for r in ranges:
        r = r.strip()
        if not r: continue
        if '-' in r:
            try:
                parts = r.split('-')
                start, end = int(parts[0]), int(parts[1])
                if start > end:
                    print_error(f"Неверный диапазон '{r}'.")
                    continue
                folder_numbers.update(range(start, end + 1))
            except ValueError:
                print_error(f"Неверный формат диапазона '{r}'.")
        else:
            try:
                folder_numbers.add(int(r))
            except ValueError:
                print_error(f"Неверный формат числа '{r}'.")
    return sorted(list(folder_numbers))


def get_number_from_string(text):
    """Извлекает первое число из строки для сортировки."""
    match = re.search(r'\d+', text)
    return int(match.group()) if match else float('inf')


def generate_range_string(processed_numbers):
    """Генерирует строку диапазона (3550-3552;3560)."""
    if not processed_numbers:
        return "NoRange"

    processed_numbers = sorted(list(set(processed_numbers)))
    range_parts = []
    current_range = [processed_numbers[0]]

    for i in range(1, len(processed_numbers)):
        if processed_numbers[i] == processed_numbers[i - 1] + 1:
            current_range.append(processed_numbers[i])
        else:
            if len(current_range) > 1:
                range_parts.append(f"{current_range[0]}-{current_range[-1]}")
            else:
                range_parts.append(str(current_range[0]))
            current_range = [processed_numbers[i]]

    if len(current_range) > 1:
        range_parts.append(f"{current_range[0]}-{current_range[-1]}")
    else:
        range_parts.append(str(current_range[0]))

    return ';'.join(range_parts)


def save_merged_pdf(merger, save_path, file_name):
    """Сохраняет PDF и обрабатывает ошибки."""
    full_path = os.path.join(save_path, file_name)
    try:
        if not os.path.exists(save_path):
            os.makedirs(save_path)

        print(f"Сохранение: {file_name} ...")
        with open(full_path, 'wb') as f_out:
            merger.write(f_out)
        merger.close()
        print(f"✅ Готово!")
    except Exception as e:
        print_error(f"Ошибка при сохранении: {e}")


def normalize_gtd_number(value):
    """Приводит номер ДТ к единому виду для надежного сопоставления."""
    if value is None:
        return ""
    raw_value = str(value).strip()
    if not raw_value:
        return ""

    # Нормализуем типичные варианты:
    # GTD_10702070_120526_5176014.pdf
    # 10702070-120526-5176014
    # 10702070/120526/5176014
    # и строки с дополнительным текстом.
    normalized = raw_value.upper().replace("'", "").replace('"', "")
    normalized = re.sub(r"\.PDF$", "", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"^GTD[_\-\s:/\\]*", "", normalized, flags=re.IGNORECASE)
    normalized = normalized.strip()

    # Предпочтительный путь: три числовых блока "код_дата_номер".
    match = re.search(r"(\d{8})\D*(\d{6})\D*(\d+)", normalized)
    if match:
        return f"{match.group(1)}_{match.group(2)}_{match.group(3)}"

    # Резерв: берем первые три числовых блока, если шаблон выше не пойман.
    digit_parts = re.findall(r"\d+", normalized)
    if len(digit_parts) >= 3:
        return f"{digit_parts[0]}_{digit_parts[1]}_{digit_parts[2]}"

    # Последний резерв: унифицируем любые разделители.
    normalized = re.sub(r"[^\dA-Z]+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized


def _parse_date_text(value):
    """Пробует распознать текстовую дату без изменения исходного значения."""
    date_formats = (
        "%d.%m.%Y",
        "%d.%m.%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
    )
    for fmt in date_formats:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def get_release_date_sort_key(value):
    """Формирует ключ сортировки по дате выпуска (от ранней к поздней)."""
    if isinstance(value, datetime):
        return (0, value.date().toordinal())
    if isinstance(value, date):
        return (0, value.toordinal())
    if isinstance(value, (int, float)):
        return (0, float(value))
    if value is None:
        return (2, "")

    value_text = str(value).strip()
    if not value_text:
        return (2, "")

    parsed_date = _parse_date_text(value_text)
    if parsed_date is not None:
        return (0, parsed_date.toordinal())

    return (1, value_text)


def load_release_dates_from_sorting_sheet():
    """Читает Sorting sheet.xlsx и возвращает словарь: номер ДТ -> дата выпуска."""
    if not os.path.exists(SORTING_SHEET_FILE):
        print_error('Файл "Sorting sheet.xlsx" не найден рядом со скриптом.')
        return None

    try:
        from openpyxl import load_workbook
    except ImportError:
        print_error('Не найден модуль openpyxl. Выполните: pip install -r requirements.txt')
        return None

    workbook = None
    try:
        workbook = load_workbook(SORTING_SHEET_FILE, data_only=True, read_only=True)
        if "TOTAL" not in workbook.sheetnames:
            print_error('В файле "Sorting sheet.xlsx" отсутствует лист "TOTAL".')
            return None

        sheet = workbook["TOTAL"]
        release_dates = {}
        duplicated_numbers = []

        for _, release_date, document_number in sheet.iter_rows(
            min_row=2, min_col=1, max_col=3, values_only=True
        ):
            normalized_number = normalize_gtd_number(document_number)
            if not normalized_number:
                continue

            if normalized_number in release_dates:
                duplicated_numbers.append(normalized_number)
                continue

            release_dates[normalized_number] = {
                "release_date": release_date,
                "release_key": get_release_date_sort_key(release_date),
            }

        if duplicated_numbers:
            duplicate_list = ", ".join(sorted(set(duplicated_numbers)))
            print_error(
                f'В файле "Sorting sheet.xlsx" есть дубли номеров ДТ: {duplicate_list}. '
                "Для дублей используется первое найденное значение."
            )

        if not release_dates:
            print_error('В файле "Sorting sheet.xlsx" не найдено номеров ДТ в колонке C.')
            return None

        return release_dates
    except Exception as e:
        print_error(f'Ошибка чтения файла "Sorting sheet.xlsx": {e}')
        return None
    finally:
        if workbook is not None:
            workbook.close()


# ==========================================
# ЛОГИКА 1: BindingInvSpec (Инвойсы и Спецификации)
# ==========================================
def process_inv_spec(source_path, save_path, valid_folders):
    print("\n[Выполняется: Инвойсы и Спецификации]")

    def get_invoice_num(fname):
        match = re.search(r'Invoice (\d+)', fname, re.IGNORECASE)
        return int(match.group(1)) if match else float('inf')

    all_invoice_pdfs = []
    processed_folders = []

    all_folders = sorted(os.listdir(source_path), key=get_number_from_string)

    for folder_name in all_folders:
        folder_path = os.path.join(source_path, folder_name)
        if not os.path.isdir(folder_path): continue

        f_num = get_number_from_string(folder_name)
        if f_num in valid_folders:
            found_in_folder = False
            for file_name in os.listdir(folder_path):
                if "invoice" in file_name.lower() and file_name.lower().endswith(".pdf"):
                    all_invoice_pdfs.append(os.path.join(folder_path, file_name))
                    found_in_folder = True

            if found_in_folder:
                processed_folders.append(f_num)

    if not all_invoice_pdfs:
        print_error("Файлы Invoice не найдены.")
        return

    all_invoice_pdfs.sort(key=lambda x: get_invoice_num(os.path.basename(x)))

    range_str = generate_range_string(processed_folders)
    output_name = f"Inv. + Spec. {range_str} {len(all_invoice_pdfs)} pcs..pdf"

    merger = PdfMerger()
    for pdf in all_invoice_pdfs:
        try:
            merger.append(pdf)
        except Exception as e:
            print_error(f"Ошибка с файлом {pdf}: {e}")

    save_merged_pdf(merger, save_path, output_name)


# ==========================================
# ЛОГИКА 2: BindingGTDESD (Декларации и ЭСД)
# ==========================================
def process_gtd_esd(source_path, save_path, valid_folders):
    print("\n[Выполняется: Декларации и ЭСД]")
    processed_folders = []
    all_pdfs = []

    all_folders = sorted(os.listdir(source_path), key=get_number_from_string)

    for folder_name in all_folders:
        folder_path = os.path.join(source_path, folder_name)
        if not os.path.isdir(folder_path): continue

        f_num = get_number_from_string(folder_name)
        if f_num in valid_folders:
            gtd_files = []
            esd_files = []

            for file_name in os.listdir(folder_path):
                if not file_name.lower().endswith(".pdf"): continue
                if file_name.startswith("GTD_"):
                    gtd_files.append(os.path.join(folder_path, file_name))
                elif file_name.count('-') == 4:
                    esd_files.append(os.path.join(folder_path, file_name))

            if gtd_files and esd_files:
                processed_folders.append(f_num)
                all_pdfs.extend(sorted(gtd_files)[:1])
                all_pdfs.extend(sorted(esd_files)[:1])
            else:
                print_error(f"Папка {folder_name} пропущена: некомплект.")

    if not all_pdfs:
        print_error("Не найдено пар GTD+ESD.")
        return

    range_str = generate_range_string(processed_folders)
    output_name = f"GTD+ЭСД {range_str} {len(processed_folders)} pcs..pdf"

    merger = PdfMerger()
    for pdf in all_pdfs:
        merger.append(pdf)

    save_merged_pdf(merger, save_path, output_name)


# ==========================================
# ЛОГИКА 3: BindingGTDInvSpec (GTD + Invoice + Spec)
# ==========================================
def process_gtd_inv_spec(source_path, save_path, valid_folders):
    print("\n[Выполняется: Декларации, Инвойсы и Спецификации]")
    release_dates_by_gtd = load_release_dates_from_sorting_sheet()
    if release_dates_by_gtd is None:
        return

    valid_pairs = []
    processed_folders_set = set()
    missing_gtd_numbers = []

    all_folders = sorted(os.listdir(source_path), key=get_number_from_string)

    for folder_name in all_folders:
        folder_path = os.path.join(source_path, folder_name)
        if not os.path.isdir(folder_path): continue

        f_num = get_number_from_string(folder_name)
        if f_num in valid_folders:
            gtd_files = []
            inv_files = []

            for file_name in sorted(os.listdir(folder_path)):
                lower_name = file_name.lower()
                if not lower_name.endswith(".pdf"): continue

                if lower_name.startswith("gtd_"):
                    gtd_files.append(os.path.join(folder_path, file_name))
                elif "invoice" in lower_name:
                    inv_files.append(os.path.join(folder_path, file_name))

            gtd_path = gtd_files[0] if gtd_files else None
            inv_path = inv_files[0] if inv_files else None

            if gtd_path and inv_path:
                normalized_gtd = normalize_gtd_number(os.path.basename(gtd_path))
                release_entry = release_dates_by_gtd.get(normalized_gtd)

                if release_entry is None:
                    missing_gtd_numbers.append(normalized_gtd or os.path.basename(gtd_path))
                    continue

                sort_key = (release_entry["release_key"], normalized_gtd)
                valid_pairs.append({
                    'sort_key': sort_key,
                    'gtd': gtd_path,
                    'inv': inv_path
                })
                processed_folders_set.add(f_num)

            elif gtd_path and not inv_path:
                print_error(f"Папка {folder_name}: Найден GTD, но нет Invoice! (Пропущено)")

            elif inv_path and not gtd_path:
                print_error(f"Папка {folder_name}: Найден Invoice, но нет GTD! (Пропущено)")

    if missing_gtd_numbers:
        missing_list = ", ".join(sorted(set(missing_gtd_numbers)))
        print_error(f'В файле "Sorting sheet.xlsx" отсутствуют номера ДТ: {missing_list}.')
        print_error('Обновите файл "Sorting sheet.xlsx" и повторите запуск.')
        return

    if not valid_pairs:
        print_error("Файлы для скрепления не найдены (или возникли ошибки комплектности).")
        return

    valid_pairs.sort(key=lambda x: x["sort_key"])

    files_to_merge = []
    for pair in valid_pairs:
        files_to_merge.append(pair['gtd'])
        files_to_merge.append(pair['inv'])

    range_str = generate_range_string(list(processed_folders_set))
    output_name = f"GTD+Inv. + Spec. {range_str} {len(processed_folders_set)} pcs..pdf"

    merger = PdfMerger()
    for pdf in files_to_merge:
        merger.append(pdf)

    save_merged_pdf(merger, save_path, output_name)


# ==========================================
# ЛОГИКА 4: BindingGTD (Только Декларации)
# ==========================================
def process_gtd_only(source_path, save_path, valid_folders):
    print("\n[Выполняется: Только Декларации (GTD)]")
    processed_folders = []
    all_pdfs = []

    all_folders = sorted(os.listdir(source_path), key=get_number_from_string)

    for folder_name in all_folders:
        folder_path = os.path.join(source_path, folder_name)
        if not os.path.isdir(folder_path): continue

        f_num = get_number_from_string(folder_name)
        if f_num in valid_folders:
            gtd_files = []
            for file_name in os.listdir(folder_path):
                if file_name.lower().endswith(".pdf") and file_name.startswith("GTD_"):
                    gtd_files.append(os.path.join(folder_path, file_name))

            if gtd_files:
                processed_folders.append(f_num)
                all_pdfs.extend(sorted(gtd_files)[:1])

    if not all_pdfs:
        print_error("GTD файлы не найдены.")
        return

    range_str = generate_range_string(processed_folders)
    output_name = f"GTD {range_str} {len(processed_folders)} pcs..pdf"

    merger = PdfMerger()
    for pdf in all_pdfs:
        merger.append(pdf)

    save_merged_pdf(merger, save_path, output_name)


# ==========================================
# ЛОГИКА 5: Ж/Д Накладные (Railway)
# ==========================================
def process_railway():
    print("\n[Выполняется: Ж/Д накладные по 4 шт.]")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    source_folder = os.path.join(script_dir, "Railway")
    save_folder = os.path.join(script_dir, "Merged Railway")

    if not os.path.exists(source_folder):
        print_error(f"Папка Railway не найдена по пути: {source_folder}")
        print("Создайте папку 'Railway' рядом со скриптом.")
        return

    files = [f for f in os.listdir(source_folder) if f.lower().endswith('.pdf')]
    if not files:
        print_error("В папке Railway нет PDF файлов.")
        return

    files.sort(key=get_number_from_string)

    chunk_size = 4
    chunks = [files[i:i + chunk_size] for i in range(0, len(files), chunk_size)]

    if not os.path.exists(save_folder):
        os.makedirs(save_folder)

    for chunk in chunks:
        merger = PdfMerger()
        file_numbers = []

        print(f"Скрепляю ({len(chunk)} шт): {chunk[0]} ... {chunk[-1]}")

        for fname in chunk:
            full_path = os.path.join(source_folder, fname)
            merger.append(full_path)
            file_numbers.append(get_number_from_string(fname))

        range_str = generate_range_string(file_numbers)
        output_name = f"Railway {range_str} {len(chunk)} pcs..pdf"

        save_merged_pdf(merger, save_folder, output_name)

    print(f"\n✅ Все файлы обработаны. Сохранено в: {save_folder}")


# ==========================================
# ЛОГИКА TEMP (Папка Temp)
# ==========================================
def process_temp_folder():
    print("\n[Выполняется: Скрепление из папки Temp]")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    temp_folder = os.path.join(script_dir, "Temp")
    combined_folder = os.path.join(script_dir, "Combined")

    if not os.path.exists(temp_folder):
        print_error("Папка Temp не найдена.")
        return

    def extract_temp_number(filename):
        match = re.match(r"^(\d+),", filename)
        return int(match.group(1)) if match else float('inf')

    pdf_files = [f for f in os.listdir(temp_folder) if f.lower().endswith(".pdf")]
    sorted_pdfs = sorted(pdf_files, key=extract_temp_number)

    if not sorted_pdfs:
        print_error("В папке Temp нет PDF файлов.")
        return

    merger = PdfMerger()
    for pdf in sorted_pdfs:
        merger.append(os.path.join(temp_folder, pdf))

    if not os.path.exists(combined_folder): os.makedirs(combined_folder)

    existing = [f for f in os.listdir(combined_folder) if f.startswith("Combined") and f.endswith(".pdf")]
    next_num = 1
    if existing:
        nums = []
        for f in existing:
            m = re.search(r"Combined-(\d+)", f)
            if m: nums.append(int(m.group(1)))
        if nums: next_num = max(nums) + 1

    out_name = f"Combined-{next_num}.pdf"
    save_merged_pdf(merger, combined_folder, out_name)


# ==========================================
# МЕНЮ (STATE MACHINE)
# ==========================================

def main():
    while True:
        # --- ГЛАВНОЕ МЕНЮ ---
        print("\n" + "=" * 45)
        print(f"   {BOLD}УТИЛИТА СКРЕПЛЕНИЯ ДОКУМЕНТОВ{RESET}")
        print("=" * 45)
        print("Выберите документы для скрепления:")
        print("1. Отгрузочные документы (GTD, Invoice, ESD)")
        print("2. Документы из папки Temp")
        print("3. Ж/Д накладные из папки Railway")
        print("0. Выход")

        main_choice = input(f"\n{BOLD}Ваш выбор:{RESET} ").strip()

        if main_choice == '0':
            print("Выход из программы.")
            break

        elif main_choice == '2':
            process_temp_folder()

        elif main_choice == '3':
            process_railway()

        elif main_choice == '1':
            shipping_docs_workflow()

        else:
            print_error("Неверный ввод.")


def shipping_docs_workflow():
    """
    Машина состояний для процесса отгрузочных документов.
    """

    # ----------------------------------------
    # ИНИЦИАЛИЗАЦИЯ: Проверка сохраненных путей
    # ----------------------------------------
    source_path = ""
    save_path = ""
    valid_folders = []

    config = load_config()
    loaded_from_config = False

    if config:
        cfg_source = config.get("source_path", "")
        cfg_save = config.get("save_path", "")

        # Проверяем, существуют ли сохраненные папки
        if os.path.isdir(cfg_source) and os.path.isdir(cfg_save):
            source_path = cfg_source
            save_path = cfg_save
            loaded_from_config = True
            current_state = 'ASK_RANGE'  # Сразу переходим к диапазону
            print(f"\nℹ️  {BOLD}Используются сохраненные пути:{RESET}")
            print(f"   📁 Откуда: {source_path}")
            print(f"   💾 Куда:   {save_path}")
        else:
            current_state = 'ASK_SOURCE'
    else:
        current_state = 'ASK_SOURCE'

    while True:
        # ----------------------------------------
        # ЭТАП 1: Выбор исходной папки
        # ----------------------------------------
        if current_state == 'ASK_SOURCE':
            print("\n🟧 Шаг 1: Исходная директория")
            print("Введите путь или '0' для возврата в Главное Меню.")
            user_input = get_clean_path("Путь к папке с инвойсами", allow_menu_codes=True)

            if user_input == '0':
                return  # Возврат в main()

            if not os.path.isdir(user_input):
                print_error("Папка не существует. Попробуйте еще раз.")
                continue

            source_path = user_input
            current_state = 'ASK_SAVE_PATH'

        # ----------------------------------------
        # ЭТАП 2: Путь сохранения
        # ----------------------------------------
        elif current_state == 'ASK_SAVE_PATH':
            print("\n🟧 Шаг 2: Место сохранения")
            print("Введите путь для сохранения готового файла.")
            print("1. Возврат к выбору исходной директории")
            print("0. Возврат в главное меню")

            user_input = get_clean_path("Путь сохранения", allow_menu_codes=True)

            if user_input == '0': return
            if user_input == '1':
                current_state = 'ASK_SOURCE'
                continue

            save_path = user_input

            # Если оба пути валидны, сохраняем конфиг
            if os.path.isdir(source_path) and os.path.isdir(save_path):
                save_config(source_path, save_path)

            current_state = 'ASK_RANGE'

        # ----------------------------------------
        # ЭТАП 3: Выбор диапазона
        # ----------------------------------------
        elif current_state == 'ASK_RANGE':
            print("\n🟧 Шаг 3: Диапазон папок")
            print("Введите диапазон (например: 3550-3553,3560)")
            print("1. Изменить путь сохранения (Назад)")
            print("9. Сбросить все пути и выбрать папку заново")
            print("0. Возврат в главное меню")

            user_input = input(f"{BOLD}Диапазон или команда:{RESET} ").strip()

            if user_input == '0': return
            if user_input == '1':
                current_state = 'ASK_SAVE_PATH'
                continue
            if user_input == '9':
                current_state = 'ASK_SOURCE'
                continue

            folders = parse_folder_range(user_input)
            if not folders:
                print_error("Некорректный диапазон.")
                continue

            print(f"✔ Будут обработаны папки: {folders}")
            valid_folders = folders
            current_state = 'SELECT_TYPE'

        # ----------------------------------------
        # ЭТАП 4: Выбор типа скрепления
        # ----------------------------------------
        elif current_state == 'SELECT_TYPE':
            print("\n🟧 Шаг 4: Выбор типа скрепления")

            # Показываем короткие версии путей для наглядности
            short_source = source_path[-30:] if len(source_path) > 30 else source_path
            short_save = save_path[-30:] if len(save_path) > 30 else save_path

            print(f"Источник: ...{short_source}")
            print(f"Сохранение в: ...{short_save}")
            print("-" * 30)
            print("1. Инвойсы и Спецификации")
            print("2. Декларации и ЭСД")
            print("3. Декларации, Инвойсы и Спецификации")
            print("4. Декларации (Только GTD)")
            print("-" * 30)
            print("6. Возврат к выбору диапазона номеров")
            print("7. Изменить пути (возврат к выбору папки)")
            print("0. Возврат в главное меню")

            choice = input(f"\n{BOLD}Ваш выбор:{RESET} ").strip()

            # Навигация
            if choice == '0': return
            if choice == '6':
                current_state = 'ASK_RANGE'
                continue
            if choice == '7':
                current_state = 'ASK_SOURCE'
                continue

            # Действия
            if choice == '1':
                process_inv_spec(source_path, save_path, valid_folders)
            elif choice == '2':
                process_gtd_esd(source_path, save_path, valid_folders)
            elif choice == '3':
                process_gtd_inv_spec(source_path, save_path, valid_folders)
            elif choice == '4':
                process_gtd_only(source_path, save_path, valid_folders)
            else:
                print_error("Неверный выбор.")
                time.sleep(1)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nПрограмма остановлена.")
